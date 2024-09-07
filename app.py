from flask import Flask, request, send_file, render_template
import pandas as pd
import os
import tempfile
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Font

app = Flask(__name__)

def detect_headers(df):
    """
    Function to detect the header row by searching for specific keywords.
    Returns the row index where the headers are found.
    """
    for i, row in df.iterrows():
        # Check if certain expected keywords (like 'S.No', 'DATE', etc.) are in the row
        if any(keyword in str(cell).upper() for cell in row for keyword in ['S.NO', 'DATE']):
            return i
    # If no header is found, assume the first row is the header
    return 0

def create_consolidated_mapping(df, temp_dir, sheet_name):
    """
    Function to create a consolidated doctor-patient mapping for a given DataFrame.
    Returns the file path where the consolidated mapping Excel is saved.
    """
    grouped = df.groupby('REF. DOCTOR')
    
    # Create an Excel file to store the consolidated mapping
    consolidated_filename = f"{sheet_name}_doctor.xlsx"
    consolidated_filepath = os.path.join(temp_dir, consolidated_filename)
    
    # Create a new workbook and a writer object
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Initialize the row index for writing data
    row_idx = 1
    
    for doctor, group in grouped:
        # Filter out rows where 'REF. DOCTOR' is NaN
        group = group.dropna(subset=['REF. DOCTOR'])
        if not group.empty:
            # Write doctor name as a section header with bold formatting
            doctor_cell = ws.cell(row=row_idx, column=1, value=doctor)
            doctor_cell.font = Font(bold=True)
            row_idx += 1
            
            # Drop the 'REF. DOCTOR' column
            group = group.drop(columns=['REF. DOCTOR'])
            
            # Write DataFrame to the Excel file
            for r_idx, row in group.iterrows():
                for c_idx, value in enumerate(row):
                    ws.cell(row=row_idx, column=c_idx+1, value=value)
                row_idx += 1
                
            row_idx += 1  # Add an empty row for separation
    
    # Save the workbook
    wb.save(consolidated_filepath)
    
    return consolidated_filepath

@app.route('/', methods=['GET'])
def home():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return "No file part", 400

    file = request.files['file']
    
    if file.filename == '':
        return "No selected file", 400
    
    if file:
        # Create a temporary directory to store the CSV files
        with tempfile.TemporaryDirectory() as temp_dir:
            # Save the uploaded file
            xlsx_filepath = os.path.join(temp_dir, file.filename)
            file.save(xlsx_filepath)
            
            # Load the Excel workbook using pandas
            xls = pd.ExcelFile(xlsx_filepath)
            mapping_file_paths = []
            # Process each sheet
            for sheet_name in xls.sheet_names:
                # Read the sheet into a DataFrame (initially without setting headers)
                df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                
                # Detect where the actual headers are located
                header_row = detect_headers(df)
                
                # Reload the DataFrame using the detected header row
                df = pd.read_excel(xls, sheet_name=sheet_name, header=header_row)
                
                # Generate a consolidated mapping Excel file for the sheet
                consolidated_mapping_file = create_consolidated_mapping(df, temp_dir, sheet_name)
                mapping_file_paths.append(consolidated_mapping_file)
            
            # Combine all the mapping files into a single zip file for download
            raw_name = file.filename.split(".")[0]
            zip_filename = f"{raw_name}.zip"
            zip_filepath = os.path.join(temp_dir, zip_filename)
            
            # Create a zip file
            with zipfile.ZipFile(zip_filepath, 'w') as zipf:
                for file_path in mapping_file_paths:
                    zipf.write(file_path, os.path.basename(file_path))
            
            # Send the zip file as a response
            return send_file(zip_filepath, as_attachment=True, download_name=zip_filename)

if __name__ == '__main__':
    app.run(debug=True)
